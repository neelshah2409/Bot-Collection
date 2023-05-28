from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "James": {
        "English": 65,
        "Physics": 78,
        "Computer": 98,
        "History": 89
    },
    "Rhea": {
        "English": 55,
        "Physics": 77,
        "Computer": 87,
        "History": 95
    },
    "Harsh": {
        "English": 100,
        "Physics": 45,
        "Computer": 75,
        "History": 92
    },
    "Suman": {
        "English": 30,
        "Physics": 25,
        "Computer": 45,
        "History": 100
    },
    "Ryan": {
        "English": 90,
        "Physics": 100,
        "Computer": 92,
        "History": 60
    }
}

wb = Workbook()
ws = wb.active
ws.title = "Student Marks"

headings = ['Name'] + list(data['James'].keys())
ws.append(headings)

for person in data:
    marks = list(data[person].values())
    ws.append([person] + marks)

for col in range(2, len(data['James']) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("StudentMarks.xlsx")
